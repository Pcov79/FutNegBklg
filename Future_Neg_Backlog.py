import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

st.set_page_config(page_title="Future Negative Backlog", layout="wide")
st.title("📊 Future Negative Backlog Check")

# Upload files
billing_file = st.file_uploader("Upload Billing Plan Excel File", type="xlsx")
backlog_file = st.file_uploader("Upload Backlog Excel File", type="xlsx")
engagement_file = st.file_uploader("Upload Engagement Manager Excel File", type="xlsx")

if billing_file and backlog_file and engagement_file:
    billing_df = pd.read_excel(billing_file, engine="openpyxl")
    backlog_df = pd.read_excel(backlog_file, engine="openpyxl")
    engagement_df = pd.read_excel(engagement_file, engine="openpyxl")

    # Ensure Billing Date is datetime
    billing_df["Billing Date"] = pd.to_datetime(billing_df["Billing Date"])
    billing_df.sort_values(by=["WBS Element", "Sales Order", "Billing Date"], inplace=True)

    # Compute Backlog Exceeded Date
    exceed_date_dict = {}
    grouped = billing_df.groupby(["WBS Element", "Sales Order"])
    for (wbs, order), group in grouped:
        group = group.copy()
        group["Cumulative Billing"] = group["Billing Value"].cumsum()
        backlog_row = backlog_df[
            (backlog_df["WBS Element"] == wbs) &
            (backlog_df["Sales Order"] == order)
        ]
        if not backlog_row.empty:
            remaining_backlog = backlog_row.iloc[0]["Remaining Backlog"]
            exceeded = group[group["Cumulative Billing"] > remaining_backlog]
            if not exceeded.empty:
                exceed_date_dict[(wbs, order)] = exceeded.iloc[0]["Billing Date"].date()
            else:
                exceed_date_dict[(wbs, order)] = None

    # Summarize billing
    billing_summary = billing_df.groupby(
        ["WBS Element", "Sales Organization", "Sales Order"], as_index=False
    )["Billing Value"].sum()

    # Summarize backlog
    backlog_summary = backlog_df.groupby(
        ["WBS Element", "Sales Organization", "Sales Order"], as_index=False
    )[
        ["Remaining Backlog", "Measurement customer Name 1"]
    ].first()

    # Merge billing and backlog
    merged_df = pd.merge(
        billing_summary,
        backlog_summary,
        on=["WBS Element", "Sales Organization", "Sales Order"],
        how="left"
    )

    # Calculate Delta Backlog
    merged_df["Delta Backlog"] = (merged_df["Remaining Backlog"] - merged_df["Billing Value"]).round(2)

    # Add Backlog Exceeded Date
    merged_df["Backlog Exceeded Date"] = merged_df.apply(
        lambda row: exceed_date_dict.get((row["WBS Element"], row["Sales Order"]), None),
        axis=1
    )

    # Add Days Left column
    today = datetime.today().date()
    merged_df["Days Left"] = merged_df["Backlog Exceeded Date"].apply(
        lambda d: (d - today).days if pd.notnull(d) else None
    )

    # Merge with engagement manager
    engagement_df = engagement_df[["Sales Document", "Eng Mgr - First name", "Eng Mgr - Last name"]]
    merged_df = pd.merge(
        merged_df,
        engagement_df,
        left_on="Sales Order",
        right_on="Sales Document",
        how="left"
    ).drop(columns=["Sales Document"])

    # Reorder columns
    ordered_columns = [
        "Sales Organization", "Sales Order", "Measurement customer Name 1", "WBS Element",
        "Billing Value", "Remaining Backlog", "Delta Backlog",
        "Backlog Exceeded Date", "Days Left", "Eng Mgr - First name", "Eng Mgr - Last name"
    ]
    merged_df = merged_df[ordered_columns]

    # Save to Excel with formatting
    output = BytesIO()
    merged_df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    # Highlight negative values
    wb = load_workbook(output)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    delta_col_idx = header.index("Delta Backlog") + 1
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=delta_col_idx, max_col=delta_col_idx):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.fill = yellow_fill

    # Save final workbook
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.download_button("Download Result Excel", data=final_output, file_name="negative_backlog_analysis.xlsx")
    st.dataframe(merged_df)
